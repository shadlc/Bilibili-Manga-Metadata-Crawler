"""Bilibili-Manga-Metadata-Crawler"""

import os
import csv
import sys
import json
import uuid
import time
import random
import hashlib
import argparse
import traceback
import tkinter as tk
from threading import Lock
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import psutil
import urllib3
import requests
from tqdm import tqdm
from colorama import Fore
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

class ArgumentParser(argparse.ArgumentParser):
    """参数类"""
    def error(self, message):
        translation_map = {
            'not allowed with argument': '无法与此参数一起使用',
            'is required': '至少需要输入一个',
            'one of the arguments': '以下参数',
            'invalid choice:': '无效的选项: ',
            'expected one argument': '需要输入一个参数',
            'unrecognized arguments': '无法识别的参数',
            'show this help message and exit': '展示这条帮助并退出',
            'invalid int value: ': '不是正确的数字: ',
            'argument ': '参数 ',
        }
        for eng, chs in translation_map.items():
            if eng in message:
                message = message.replace(eng, chs)

        self.print_usage(sys.stderr)
        print(f'{self.prog}: {message}', file=sys.stderr)
        sys.exit(2)

    def id_list(self, value):
        """解析 id 参数, 支持空格或逗号分隔"""
        return [x.strip() for x in value.replace(',', ' ').split() if x.strip()]

def parse_args():
    """参数"""
    parser = ArgumentParser(
        description='bmmc - 哔哩哔哩漫画元数据请求器',
        add_help=False
    )
    group = parser.add_mutually_exclusive_group(required=True)
    parser.add_argument('-h', '--help', help='显示此帮助信息并退出', action='help')
    group.add_argument('-p', '--parameter', action='store_true', help='参数列表')
    parser.add_argument('-y', '--yes', action='store_true', help='自动确认, 无需提示')
    parser.add_argument('-d', '--detail', action='store_true', help=f'{Fore.RED}请求漫画详情页(确保数据完整){Fore.RESET}')
    parser.add_argument('-b', '--bonus', action='store_true', help=f'{Fore.RED}保存特典信息{Fore.RESET}')
    group.add_argument('-t', '--type', help='获取不同分类页面的漫画数据，详情参考参数列表')
    group.add_argument('-i', '--id', help='输入一个或多个ID, 使用空格或逗号分隔', type=parser.id_list)
    parser.add_argument('-s', '--style', help='分类页中选择风格，详情参考参数列表', type=int, default=-1)
    parser.add_argument('-a', '--area', help='分类页中选择地区，详情参考参数列表', type=int, default=-1)
    parser.add_argument('-u', '--status', help='分类页中选择状态，详情参考参数列表', type=int, default=-1)
    parser.add_argument('-o', '--order', help='分类页中选择排序方式，详情参考参数列表', type=int, default=0)
    parser.add_argument('-c', '--price', help='分类页中选择收费方式，详情参考参数列表', type=int, default=-1)
    parser.add_argument('-e', '--special', help='分类页中选择特殊分类，详情参考参数列表', type=int, default=0)
    parser.add_argument('-r', '--rank', help='排行页中选择排行类型，详情参考参数列表', type=int, default=0)
    parser.add_argument('--sdate', help='更新推荐页中选择开始日期', default=time.strftime("%Y-%m-%d", time.localtime()))
    parser.add_argument('--edate', help='更新推荐页中选择结束日期', default=time.strftime("%Y-%m-%d", time.localtime()))
    group.add_argument('-I', '--input', help='指定读取数据的文件, 支持json、csv、xlsx')
    parser.add_argument('-O', '--output', help='指定输出文件名以及格式, 支持json、csv、xlsx', default="metadata.json")
    parser.add_argument('-w', '--workers', help='并发线程数量', type=int, default=4)
    parser.add_argument('-D', '--delay', help='如果是单线程作业, 每个请求间隔(单位: 毫秒)', type=int, default=0)
    parser.add_argument('-H', '--headers', help='请求头文件(json格式), 可包含Cookie')
    parser.add_argument('-S', '--page_size', help='指定多页请求每页数量', type=int, default=100)
    parser.add_argument('-P', '--page_num', help='指定第几页', type=int)


    args = parser.parse_args()
    if (datetime.strptime(args.edate, "%Y-%m-%d") - datetime.strptime(args.sdate, "%Y-%m-%d")).days < 0:
        parser.error(f"开始日期需要在结束日期之前: --sdate={args.sdate} > --edate={args.edate}")
    if args.type and args.type not in ("classify", "update", "ranking", "home_feed", "favorite", "buy"):
        parser.error("请输入正确的 --type")

    return args

class Crawler:
    """请求类"""
    def __init__(self, args: argparse.Namespace):
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        self.args = args
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36",
            "Cookie": f"buvid3={uuid.uuid4()}infoc;"
        }
        try:
            if self.args.headers:
                self.headers = self.headers | json.load(open(self.args.headers, encoding="utf-8"))
        except json.JSONDecodeError as e:
            raise RuntimeError(f" {Fore.RED}--headers={self.args.headers} 不是正确的json文件{Fore.RESET}") from e
        self.req_type = {"classify": "分类页", "update": "更新推荐页", "ranking": "排行页", "home_feed": "主页信息流", "favorite": "我的追漫", "buy": "已购漫画"}
        self.classify_dict = {
            "special": {1: "Vomic有声漫", 2: "番剧原作",3: "制作人"},
            "prices": {1: "免费", 3: "等就免费",4: "限时免费", 5: "通用券可用", 6: "漫读券可用", 7: "限免卡可用", 8: "打折卡可用"},
            "orders": {5: "阅读热度", 6: "弹幕热度",7: "评论热度"},
        }
        self.ranking_dict = {}

    def confirm(self, default=True):
        """确认提示"""
        if self.args.yes:
            return True

        analyze_type = self.req_type.get(self.args.type)
        prompt = f"您选择了[{analyze_type}]"
        if self.args.id:
            prompt = f"您选择了{len(self.args.id)}本漫画, 请求漫画详情速度({self.args.workers}线程{f", 间隔{self.args.delay}毫秒" if self.args.delay else ""})"
        elif self.args.type == "ranking":
            prompt += f", 排行榜为[{self.ranking_dict.get(self.args.rank, self.args.rank)}]"
        elif self.args.type == "classify":
            prompt += f", 分类为{self.parse_classify_dict()}, 每页{self.args.page_size}本漫画{f", 指定第{self.args.page_num}页" if self.args.page_num else ""}"
        elif self.args.type == "update":
            delta = (datetime.strptime(self.args.edate, "%Y-%m-%d")- datetime.strptime(self.args.sdate, "%Y-%m-%d")).days + 1
            prompt += f", 日期为[{self.args.sdate}至{self.args.edate}]共{delta}天"
        elif self.args.type == "home_feed":
            prompt += f", 每页{self.args.page_size}本漫画(注意: 信息流加载大于200本/页可能会造成重复){f", 指定第{self.args.page_num}页" if self.args.page_num else ""}"
        if self.args.detail:
            prompt += ", 以及请求漫画详情页"
        if self.args.bonus:
            prompt += ", 同时请求漫画特典"
        if self.args.bonus or self.args.detail:
            prompt += f", 请求速度为({self.args.workers}线程{f", 间隔{self.args.delay}毫秒" if self.args.delay else ""})"
        if self.args.output:
            prompt += f", 保存文件为[{self.args.output}]"
        prompt += ", 是否继续？(Y/n): "
        ans = input(f"{Fore.YELLOW}{prompt}{Fore.RESET}").strip().lower()
        if not ans:
            return default
        if ans in ['y', 'yes']:
            return True
        elif ans in ['n', 'no']:
            return False
        else:
            return False

    def get(self, *args, **kwargs):
        """GET请求"""
        kwargs['verify'] = False
        return requests.get(*args, **kwargs)

    def post(self, *args, **kwargs):
        """POST请求"""
        kwargs['verify'] = False
        return requests.post(*args, **kwargs)

    def get_parameter(self) -> str:
        """获取参数列表"""
        labels = self.get_classify_label()
        if isinstance(labels, str):
            return labels
        for label in labels:
            if labels[label]:
                self.classify_dict[label] = {**{i["id"]: i["name"] for i in labels[label]}, **self.classify_dict.get(label, {})}

        rank_info: dict = self.get_ranking_page().get("rankInfo", {}).get("list", [])
        for item in rank_info:
            self.ranking_dict[item["id"]] = item["name"]

    def parse_parameter(self) -> None:
        """解析参数列表"""
        result = f"{Fore.YELLOW}以下是各分类与其对应可用参数列表{Fore.RESET}\n"
        result += " type | classify: 分类页, update: 更新推荐页, ranking: 排行页, home_feed: 主页信息流, favorite: 我的追漫, buy: 已购漫画\n"
        result += f" {Fore.CYAN}classify(分类页):{Fore.RESET}\n"
        for key in self.classify_dict:
            result += f"   {key.ljust(7)} | {", ".join([f"{id}: {name}" for id, name in self.classify_dict[key].items()])}\n"
        result += f" {Fore.CYAN}update(更新推荐页):{Fore.RESET}\n"
        result += "   sdate   | 不小于 2019-05-06\n"
        result += f"   edate   | 不大于 {datetime.today().strftime('%Y-%m-%d')}\n"
        result += f" {Fore.CYAN}ranking(排行页):{Fore.RESET}\n"
        result += f"   {"rank".ljust(7)} | {", ".join([f"{id}: {name}" for id, name in self.ranking_dict.items()])}\n"
        result += f" {Fore.CYAN}favorite(我的追漫):{Fore.RESET}\n"
        result += f"   {"order".ljust(7)} | 1: 追漫顺序, 2: 更新时间, 3: 最近阅读\n"
        return result

    def parse_classify_dict(self) -> str:
        """解析分类页风格"""
        text = ""
        if self.args.style != -1:
            text += f"[风格: {self.classify_dict.get("styles", {}).get(self.args.style, self.args.style)}]"
        if self.args.area  != -1:
            text += f"[地区: {self.classify_dict.get("areas", {}).get(self.args.area, self.args.area)}]"
        if self.args.status != -1:
            text += f"[状态: {self.classify_dict.get("status", {}).get(self.args.status, self.args.status)}]"
        if self.args.order  != 0:
            text += f"[排序方式: {self.classify_dict.get("orders", {}).get(self.args.order, self.args.order)}]"
        if self.args.price != -1:
            text += f"[收费方式: {self.classify_dict.get("prices", {}).get(self.args.price, self.args.price)}]"
        if self.args.special != 0:
            text += f"[特殊分类: {self.classify_dict.get("special", {}).get(self.args.special, self.args.special)}]"
        if text:
            return text
        else:
            return "[全部类型]"

    def get_classify_label(self) -> dict:
        """获取全部分类"""
        url = "https://manga.bilibili.com/twirp/comic.v1.Comic/AllLabel"
        try:
            response = self.post(url, headers=self.headers, timeout=5)
            response.raise_for_status()
            return response.json().get("data", {})
        except requests.exceptions.HTTPError as e:
            raise RuntimeError(f"请求错误 {e}") from e
        except requests.RequestException as e:
            raise RuntimeError(f"网络错误 {e}") from e
        except json.JSONDecodeError as e:
            raise RuntimeError(f"返回解析错误 {e}") from e

    def get_classify_page(self, style=-1, area=-1, status=-1, order=0, special=0, price=-1, page_num=1, page_size=100) -> dict:
        """获取分类页结果"""
        url = "https://manga.bilibili.com/twirp/comic.v1.Comic/ClassPage?device=h5&platform=web"
        payload = {
            "style_id": style,
            "area_id": area,
            "is_finish": status,
            "order": order,
            "special_tag": special,
            "is_free": price,
            "page_num": page_num,
            "page_size": page_size,
        }
        try:
            response = self.post(url, headers=self.headers, data=payload, timeout=5)
            response.raise_for_status()
            data = response.json().get("data", {})
            for i in data:
                i["comic_id"] = i.get("season_id")
            return data
        except requests.exceptions.HTTPError as e:
            raise RuntimeError(f"请求错误 {e}") from e
        except requests.RequestException as e:
            raise RuntimeError(f"网络错误 {e}") from e
        except json.JSONDecodeError as e:
            raise RuntimeError(f"返回解析错误 {e}") from e

    def get_update_page(self, date: str, page_num=1, page_size=100) -> dict:
        """获取推荐页结果"""
        url = "https://manga.bilibili.com/twirp/comic.v1.Comic/GetDailyPush"
        payload = {
            "date": date,
            "page_num": page_num,
            "page_size": page_size,
        }
        try:
            response = self.post(url, headers=self.headers, data=payload, timeout=5)
            response.raise_for_status()
            data = response.json().get("data", {})
            comics = data.get("list")
            for comic in comics:
                comic["date"] = date
            if comics is None:
                return []
            return comics
        except requests.exceptions.HTTPError as e:
            raise RuntimeError(f"请求错误 {e}") from e
        except requests.RequestException as e:
            raise RuntimeError(f"网络错误 {e}") from e
        except json.JSONDecodeError as e:
            raise RuntimeError(f"返回解析错误 {e}") from e

    def get_ranking_page(self, rank_type="0") -> dict:
        """获取排行页结果"""
        if rank_type is None:
            rank_type = "0"
        url = f"https://manga.bilibili.com/ranking/{rank_type}/index.pageContext.json"
        try:
            response = self.get(url, headers=self.headers, timeout=5)
            response.raise_for_status()
            data = response.json().get("data", {})
            return data
        except requests.exceptions.HTTPError as e:
            raise RuntimeError(f"请求错误 {e}") from e
        except requests.RequestException as e:
            raise RuntimeError(f"网络错误 {e}") from e
        except json.JSONDecodeError as e:
            raise RuntimeError(f"返回解析错误 {e}") from e

    def get_comic_details(self, comic_id: str) -> dict:
        """获取漫画详情页"""
        url = "https://manga.bilibili.com/twirp/comic.v1.Comic/ComicDetail?device=h5&platform=web"
        try:
            response = self.post(url, headers=self.headers, data={"comic_id": comic_id}, timeout=5)
            response.raise_for_status()
            comic = response.json().get("data", {})
            comic["comic_id"] = comic.get("id")
            last_episode = comic.get("ep_list")[0]
            comic["last_ep_id"] = last_episode["id"]
            comic["last_ep_cover"] = last_episode["cover"]
            comic["last_ep_title"] = f"{last_episode["short_title"]} {last_episode["title"]}"
            comic["last_ep_date"] = last_episode["pub_time"].split(" ")[0]
            if comic["release_time"] == "":
                comic["release_time"] = comic.get("ep_list")[-1]["pub_time"].split(" ")[0]
            else:
                comic["release_time"] = comic["release_time"].replace(".","-")
            return comic
        except requests.exceptions.HTTPError as e:
            raise RuntimeError(f"请求错误 {e}") from e
        except requests.RequestException as e:
            raise RuntimeError(f"网络错误 {e}") from e
        except json.JSONDecodeError as e:
            raise RuntimeError(f"返回解析错误 {e}") from e

    def get_comic_bonus(self, comic_id: str) -> dict:
        """获取漫画特典页"""
        if comic_id is None:
            return
        url = "https://manga.bilibili.com/twirp/comic.v1.Comic/GetComicAlbumPlus?mobi_app=android_comic&device=android&platform=android&version=5.21.0"
        try:
            response = self.post(url, headers=self.headers, data={"comic_id": comic_id}, timeout=5)
            response.raise_for_status()
            data = response.json().get("data", {}).get("list", {})
            return [comic_id, data]
        except requests.exceptions.HTTPError as e:
            raise RuntimeError(f"请求错误 {e}") from e
        except requests.RequestException as e:
            raise RuntimeError(f"网络错误 {e}") from e
        except json.JSONDecodeError as e:
            raise RuntimeError(f"返回解析错误 {e}") from e

    def get_home_feeds(self, buvid=None, page_num=1, page_size=100) -> dict:
        """获取主页信息流结果"""
        url = "https://manga.bilibili.com/twirp/comic.v1.Home/HomeFeed"
        if buvid is None:
            mac = ':'.join(''.join(random.choices('0123456789ABCDEF', k=2)) for _ in range(6))
            h = hashlib.md5(mac.replace(':', '').replace('-', '').encode()).hexdigest().upper()
            buvid = 'XX' + (h[2] + h[12] + h[22] if len(h) >= 23 else '000') + h
        url += f"?buvid={buvid}"
        payload = {
            "page_num": page_num,
            "page_size": page_size,
        }
        try:
            response = self.post(url, headers=self.headers, data=payload, timeout=5)
            response.raise_for_status()
            data = response.json().get("data", {})
            feeds = data.get("feeds")
            comics = []
            for feed in feeds:
                comics.append({
                    "comic_id": feed.get("item_id"),
                    "title": feed.get("title"),
                    "type": feed.get("type"),
                    "image": feed.get("image"),
                    **feed.get("comic_info")
                })
            if comics is None:
                return []
            return comics
        except requests.exceptions.HTTPError as e:
            raise RuntimeError(f"请求错误 {e}") from e
        except requests.RequestException as e:
            raise RuntimeError(f"网络错误 {e}") from e
        except json.JSONDecodeError as e:
            raise RuntimeError(f"返回解析错误 {e}") from e

    def get_favorite(self, page_num=1, page_size=100, order=0) -> dict:
        """获取我的追漫"""
        url = "https://manga.bilibili.com/twirp/bookshelf.v1.Bookshelf/ListFavorite?device=pc&platform=web"
        payload = {
            "page_num": page_num,
            "page_size": page_size,
            "order": order,
        }
        try:
            response = self.post(url, headers=self.headers, data=payload, timeout=5)
            response.raise_for_status()
            comics = response.json().get("data", {})
            if comics is None:
                return []
            for comic in comics:
                comic["is_finish"] = {2: 0, 3: 1}.get(comic.get("status"))
                comic["total"] = comic.get("ord_count")
                comic["last_ep_title"] = comic.get("latest_ep_short_title")
                comic["last_ep_date"] = comic.get("last_ep_publish_time", "").split(" ")[0]
            return comics
        except requests.exceptions.HTTPError as e:
            if response.status_code == 401:
                print(f"{Fore.RED}账号授权失败, 您的Cookie未填写或已失效, 请使用参数--headers导入正确填写Cookie的json文件!{Fore.RESET}")
                return []
            raise RuntimeError(f"请求错误 {e}") from e
        except requests.RequestException as e:
            raise RuntimeError(f"网络错误 {e}") from e
        except json.JSONDecodeError as e:
            raise RuntimeError(f"返回解析错误 {e}") from e

    def get_buy_comics(self, page_num=1, page_size=100) -> dict:
        """获取已购漫画"""
        url = "https://manga.bilibili.com/twirp/user.v1.User/GetAutoBuyComics?device=pc&platform=web"
        payload = {
            "page_num": page_num,
            "page_size": page_size,
        }
        try:
            response = self.post(url, headers=self.headers, data=payload, timeout=5)
            response.raise_for_status()
            comics = response.json().get("data", {})
            if comics is None:
                return []
            for comic in comics:
                comic["title"] = comic.get("comic_title")
                comic["total"] = comic.get("last_ord")
                comic["last_ep_title"] = comic.get("last_short_title")
                comic["last_ep_date"] = comic.get("ctime", "").split(" ")[0]
            return comics
        except requests.exceptions.HTTPError as e:
            if response.status_code == 401:
                print(f"{Fore.RED}账号授权失败, 您的Cookie未填写或已失效, 请使用参数--headers导入正确填写Cookie的json文件!{Fore.RESET}")
                return []
            raise RuntimeError(f"请求错误 {e}") from e
        except requests.RequestException as e:
            raise RuntimeError(f"网络错误 {e}") from e
        except json.JSONDecodeError as e:
            raise RuntimeError(f"返回解析错误 {e}") from e

    def get_comics_details(self, comic_id_list):
        """批量获取漫画详情"""
        if self.args.workers == 1:
            concurrent = False
        else:
            concurrent = True
        task_list = []
        for comic_id in comic_id_list:
            task_list.append(lambda comic_id=comic_id: self.get_comic_details(comic_id))
        tr = TaskRunner(
            task_list,
            concurrent=concurrent,
            max_workers=self.args.workers,
            delay=self.args.delay,
            title="批量请求漫画详情"
        )
        tr.start()
        return tr.results

    def get_classify_page_all(self):
        """获取全部分类页"""
        data = []
        page = [{}]
        page_num = self.args.page_num if self.args.page_num else 1
        with tqdm(desc=f"分类页加载中({self.args.page_size}本/页)", unit="页") as process_bar:
            while len(page) > 0:
                page = self.get_classify_page(
                    style=self.args.style,
                    area=self.args.area,
                    status=self.args.status,
                    order=self.args.order,
                    special=self.args.special,
                    price=self.args.price,
                    page_size=self.args.page_size,
                    page_num=page_num
                )
                if len(page) == 0:
                    break
                data += page
                page_num += 1
                process_bar.update(1)
                if self.args.page_num:
                    break
        tqdm.write(f"{Fore.GREEN}加载完毕, 共{len(data)}本漫画{Fore.RESET}")
        return data

    def get_update_page_all(self):
        """批量获取更新推荐页"""
        if self.args.workers == 1:
            concurrent = False
        else:
            concurrent = True
        task_list = []

        start = datetime.strptime(self.args.sdate, "%Y-%m-%d")
        end = datetime.strptime(self.args.edate, "%Y-%m-%d")
        current = start
        while current <= end:
            task_list.append(lambda date=current.strftime("%Y-%m-%d"): self.get_update_page(date))
            current += timedelta(days=1)
        tr = TaskRunner(
            task_list,
            concurrent=concurrent,
            max_workers=self.args.workers,
            delay=self.args.delay,
            title="批量获取更新推荐页",
            unit="页"
        )
        tr.start()
        comics = []
        for daily_comics in tr.results:
            comics += daily_comics
        comics = sorted(comics, key=lambda x: x["date"])
        return comics

    def get_home_feeds_all(self) -> dict:
        """获取全部主页信息流"""
        data = []
        page = [{}]
        page_num = self.args.page_num if self.args.page_num else 1
        mac = ':'.join(''.join(random.choices('0123456789ABCDEF', k=2)) for _ in range(6))
        h = hashlib.md5(mac.replace(':', '').replace('-', '').encode()).hexdigest().upper()
        buvid = 'XX' + (h[2] + h[12] + h[22] if len(h) >= 23 else '000') + h
        print(f"{Fore.YELLOW}本次主页信息流使用buvid={buvid}{Fore.RESET}")
        with tqdm(desc=f"主页信息流加载中({self.args.page_size}本/页)", unit="页") as process_bar:
            while len(page) > 0:
                page = self.get_home_feeds(
                    buvid=buvid,
                    page_size=self.args.page_size,
                    page_num=page_num
                )
                if len(page) == 0:
                    break
                data += page
                page_num += 1
                process_bar.update(1)
                if self.args.page_num:
                    break
        tqdm.write(f"{Fore.GREEN}主页信息流加载完毕, 共{len(data)}本漫画{Fore.RESET}")
        return data

    def get_comic_bonus_all(self, comics: list) -> dict:
        """批量获取漫画特典页"""
        if self.args.workers == 1:
            concurrent = False
        else:
            concurrent = True
        task_list = []
        for comic in comics:
            task_list.append(lambda comic_id=comic.get("comic_id"): self.get_comic_bonus(comic_id))
        tr = TaskRunner(
            task_list,
            concurrent=concurrent,
            max_workers=self.args.workers,
            delay=self.args.delay,
            title="批量请求漫画特典",
            is_dict=True
        )
        tr.start()
        for comic in comics:
            comic_id = comic.get("comic_id")
            if comic_id in tr.results:
                bonus = tr.results[comic_id]
                comic["bonus"] = bonus
                comic["bonus_total"] = len(bonus)
                if len(bonus) == 0:
                    continue
                comic["last_bonus_title"] = max(bonus, key=lambda x: x["item"]["online_time"])["item"]["title"]
                comic["last_bonus_date"] = max(bonus, key=lambda x: x["item"]["online_time"])["item"]["online_time"].split(" ")[0]
                future_bonus = [item for item in bonus if datetime.strptime(item["item"]["offline_time"].split(" ")[0], '%Y-%m-%d') > datetime.today()]
                if len(future_bonus) == 0:
                    continue
                comic["recently_lock_bonus_title"] = min(future_bonus, key=lambda x: x["item"]["offline_time"])["item"]["title"]
                comic["recently_lock_bonus_date"] = min(future_bonus, key=lambda x: x["item"]["offline_time"])["item"]["offline_time"].split(" ")[0]
        return comics

    def get_favorite_all(self) -> dict:
        """获取全部我的追漫"""
        comics = self.get_favorite("1", "1000", self.args.order)
        return comics

    def get_buy_comics_all(self) -> dict:
        """获取全部已购漫画"""
        comics = self.get_buy_comics("1", "1000")
        return comics

class TaskRunner:
    """任务类"""
    def __init__(self, tasks, concurrent=False, delay=0,
                 max_workers=4, retries=0, retry_delay=1,
                 title="", unit="个", is_dict=False):
        """
        :param tasks: List[Any] 要处理的任务列表
        :param concurrent: bool 是否并发执行任务
        :param delay: float 顺序模式下任务之间的延迟（毫秒）
        :param max_workers: int 并发线程数
        :param retries: int 每个任务失败后的最大重试次数
        :param retry_delay: float 每次重试之间的等待时间（秒）
        :param title: str 展示的进度描述
        :param unit: str 进度单位
        :param is_dict: bool 是否按字典方式处理结果
        """
        self.tasks = tasks
        self.concurrent = concurrent
        self.delay = delay
        self.max_workers = max_workers
        self.retries = retries
        self.retry_delay = retry_delay
        self.title = title
        self.unit = unit
        self.is_dict = is_dict

        self.results = {} if is_dict else []
        self.total = len(tasks)
        self._lock = Lock()

    def _execute_task(self, task):
        result = None
        for attempt in range(1, self.retries + 2):  # 尝试次数 = 重试次数 + 第一次
            try:
                result = task()
                break
            except Exception as e:
                print(f"{e}")
                if attempt <= self.retries:
                    time.sleep(self.retry_delay / 1000)
        return result

    def start(self):
        if self.concurrent:
            self._run_concurrent()
        else:
            self._run_sequential()

    def _run_sequential(self):
        for task in tqdm(self.tasks, desc=f"{self.title}中", unit=self.unit):
            result = self._execute_task(task)
            with self._lock:
                if result is not None:
                    if self.is_dict:
                        self.results[result[0]] = result[1]
                    else:
                        self.results.append(result)
            if self.delay > 0:
                time.sleep(self.delay / 1000)

    def _run_concurrent(self):
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_task = {executor.submit(self._execute_task, task): task for task in self.tasks}
            for future in tqdm(as_completed(future_to_task), total=len(self.tasks), desc=f"{self.title}中", unit=self.unit):
                result = future.result()
                with self._lock:
                    if result is not None:
                        if self.is_dict:
                            self.results[result[0]] = result[1]
                        else:
                            self.results.append(result)

class Document:
    """文件处理类"""
    def __init__(self, args: argparse.Namespace):
        self.args = args
        self.type: str
        self.field_map = {
            "comic_id": "ID",
            "title": "漫画名",
            "authors": "作者",
            "info": "更新信息",
            "is_finish": "状态",
            "total": "章节数",
            "bonus_total": "特典数",
            "last_ep_title": "最新章节标题",
            "last_ep_date": "最新章节更新日期",
            "last_bonus_title": "最新特典标题",
            "last_bonus_date": "最新特典日期",
            "recently_lock_bonus_title": "最近下架特典标题",
            "recently_lock_bonus_date": "最近下架特典日期",
            "renewal_time": "更新规则",
            "introduction": "简介",
            "styles": "风格",
            "tags": "标签",
            "horizontal_covers": "横版封面",
            "vertical_cover": "竖版封面",
            "square_cover": "方形封面",
            "release_time": "上架时间",
        }
        self.field_ref = "A1:U1"
        if args.detail and args.bonus:
            pass
        elif args.detail:
            self.field_map.pop("bonus_total")
            self.field_map.pop("last_bonus_title")
            self.field_map.pop("last_bonus_date")
            self.field_map.pop("recently_lock_bonus_title")
            self.field_map.pop("recently_lock_bonus_date")
            self.field_ref = "A1:P1"
        elif args.type == "classify":
            self.field_map = {
                "comic_id": "ID",
                "title": "漫画名",
                "authors": "作者",
                "info": "更新信息",
                "is_finish": "状态",
                "total": "章节数",
                "introduction": "简介",
                "styles": "风格",
                "rd_tag": "标签",
                "horizontal_covers": "横版封面",
                "vertical_cover": "竖版封面",
                "square_cover": "方形封面",
                "release_time": "上架时间",
            }
            self.field_ref = "A1:M1"
        elif args.type == "update":
            self.field_map = {
                "comic_id": "ID",
                "title": "漫画名",
                "ep_id": "最新章节ID",
                "ep_title": "最新章节名",
                "short_title": "最新章节短名",
                "comment_total": "评论数",
                "allow_wait_free": "等免",
                "styles": "风格",
                "url": "横版封面",
                "vertical_cover": "竖版封面",
                "date": "更新日期",
            }
            self.field_ref = "A1:K1"
        elif args.type == "ranking":
            self.field_map = {
                "comic_id": "ID",
                "title": "漫画名",
                "rank": "本周排行",
                "last_rank": "上周排行",
                "authors": "作者",
                "info": "更新信息",
                "is_finish": "状态",
                "last_short_title": "最新章节短标题",
                "total": "章节数",
                "fans": "追更人数",
                "styles": "风格",
                "tags": "标签",
                "vertical_cover": "竖版封面",
            }
            self.field_ref = "A1:M1"
        elif args.type == "home_feed":
            self.field_map = {
                "comic_id": "ID",
                "title": "漫画名",
                "image": "封面",
                "is_finish": "状态",
                "lastest_short_title": "最新章节短标题",
                "main_style_name": "风格",
                "score": "评分",
                "introduction": "简介",
                "evaluate": "剧情梗概",
            }
            self.field_ref = "A1:I1"
        elif args.type == "favorite":
            self.field_map = {
                "comic_id": "ID",
                "title": "漫画名",
                "authors": "作者",
                "info": "更新信息",
                "is_finish": "状态",
                "total": "章节数",
                "last_ep_title": "最新章节短标题",
                "last_ep_date": "最新章节更新日期",
                "hcover": "横版封面",
                "vcover": "竖版封面",
                "scover": "方形封面",
                "latest_read_time": "上次阅读时间",
            }
            self.field_ref = "A1:L1"
        elif args.type == "buy":
            self.field_map = {
                "comic_id": "ID",
                "title": "漫画名",
                "total": "章节数",
                "bought_ep_count": "已购章节数",
                "last_ep_title": "最新章节短标题",
                "last_ep_date": "最新章节更新日期",
                "hcover": "横版封面",
                "vcover": "竖版封面",
                "scover": "方形封面",
            }
            self.field_ref = "A1:I1"

    def load(self) -> list:
        """载入数据"""
        data: str
        ext = os.path.splitext(self.args.input)[-1].lower()
        if ext == '.json':
            with open(self.args.input, 'r', encoding='utf-8') as f:
                data = json.load(f)
        elif ext == '.csv':
            with open(self.args.input, 'r', encoding='utf-8') as f:
                data = list(csv.DictReader(f))
                for item in data:
                    item["comic_id"] = item["ID"]
        elif ext == '.xlsx':
            wb = load_workbook(self.args.input)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            data = [dict(zip(rows[0], row)) for row in rows[1:]]
            for item in data:
                item["comic_id"] = item["ID"]
        else:
            raise ValueError(f"{Fore.RED}仅支持json、csv、xlsx格式的导入读取{Fore.RESET}")
        return data
        

    def save(self, data: dict):
        """保存为文件"""
        lower_name = self.args.output.lower()
        if lower_name.endswith('.xlsx'):
            self.type = 'xlsx'
            self.xlsx(data)
        elif lower_name.endswith('.csv'):
            self.type = 'csv'
            self.csv(data)
        else:
            self.type = 'json'
            self.json(data)

    def json(self, data: dict):
        """保存为json"""
        try:
            for item in data:
                item.pop("ep_list", None)
                item.pop("styles2", None)
                item.pop("fav_comic_info", None)
                item.pop("series_info", None)
                item.pop("story_elems", None)
                item.pop("discount_marketing", None)
                item.pop("data_info", None)
                item.pop("coupon_marketing", None)
                item.pop("discount_banner", None)
            json.dump(data, open(self.args.output, mode="w",encoding="utf-8"), ensure_ascii=False,indent=4)
        except TypeError as e:
            raise RuntimeError(f"数据异常, 保存错误！ {e}") from e
        except PermissionError as e:
            raise RuntimeError(f"数据保存失败, 无写入权限！ {e}") from e

    def mapping_field(self, field: str, row: dict) -> dict:
        """处理输出字段"""
        value = row.get(field, "")
        if field == "is_finish":
            value = {0: "连载中", 1: "已完结"}.get(value, value)
        elif field == "info" and not value:
            total = row.get("total")
            if not total:
                value = ""
            elif row.get("is_finish"):
                value = f"[已完结]共{total}话"
            else:
                value = f"[连载中]至{total}话"
        elif field == "authors":
            if value and isinstance(value[0], dict):
                value = ",".join([i.get("name", "") for i in value])
            elif row.get("author_name"):
                value = ",".join(row.get("author_name"))
            elif row.get("author"):
                value = ",".join(row.get("author"))
        elif field == "styles":
            if len(value) == 0:
                value = ""
            elif isinstance(value[0], str):
                value = ",".join(value)
            else:
                value = ",".join([i.get("name", "") for i in value])
        elif field == "tags":
            if len(value) == 0:
                value = ""
            elif isinstance(value[0], str):
                value = ",".join(value)
            else:
                value = ",".join([i.get("name", "") for i in value])
        elif field == "horizontal_covers":
            if value and isinstance(value, list):
                value = ",".join(value)
            elif row.get("horizontal_cover"):
                value = row.get("horizontal_cover")
        elif field == "allow_wait_free":
            value = "是" if value else "否"
        return value

    def xlsx(self, data: dict):
        """保存为xlsx"""
        try:
            field_keys = list(self.field_map.keys())
            headers = [self.field_map[field] for field in field_keys]
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
            ws.auto_filter.ref = self.field_ref
            for col_num, cell in enumerate(ws[1], 1):
                column_letter = get_column_letter(col_num)
                adjusted_width = len(str(cell.value)) * 2 + 5
                ws.column_dimensions[column_letter].width = adjusted_width
            for row in data:
                row_data = []
                for field in field_keys:
                    value = self.mapping_field(field, row)
                    row_data.append(value)
                ws.append(row_data)
            for cell in ws[1]:
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True)
            wb.save(self.args.output)
        except PermissionError as e:
            raise RuntimeError(f"数据保存失败, 无写入权限！ {e}") from e

    def csv(self, data: dict, encoding="utf-8-sig"):
        """保存为csv"""
        try:
            field_keys = list(self.field_map.keys())
            headers = [self.field_map[field] for field in field_keys]
            with open(self.args.output, mode="w", newline="", encoding=encoding) as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                for row in data:
                    row_data = []
                    for field in field_keys:
                        value = self.mapping_field(field, row)
                        row_data.append(value)
                    writer.writerow(row_data)
        except PermissionError as e:
            raise RuntimeError(f"数据保存失败, 无写入权限！ {e}") from e

def is_launched_by_explorer():
    """判断是否是双击运行(父进程为 explorer)"""
    try:
        parent = psutil.Process(os.getpid()).parent().parent()
        return parent and 'explorer' in parent.name().lower()
    except psutil.AccessDenied:
        return False

def get_res_path(relative_path):
    """ 获取资源的绝对路径, 适用于开发环境和 PyInstaller 打包后的环境 """
    if hasattr(sys, '_MEIPASS'):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def run_gui():
    """GUI 模式"""
    root = tk.Tk()
    root.title("哔哩哔哩漫画元数据请求器")
    root.iconbitmap(get_res_path("BiliBili_favicon.ico"))
    root.geometry("360x100")
    label = tk.Label(root, text="请使用CLI模式, GUI模式开发中...ο(=•ω＜=)ρ⌒☆")
    label.pack(pady=20, padx=20)
    root.mainloop()

def run_cli(args: argparse.Namespace, cl: Crawler, dm: Document):
    """CLI 模式"""
    cl.get_parameter()
    if args.parameter:
        print(cl.parse_parameter())
    if args.input:
        args.id = [item['comic_id'] for item in dm.load()]
        if cl.confirm():
            comics = cl.get_comics_details(args.id)
            dm.save(comics)
            tqdm.write(f"{Fore.GREEN}自定义漫画ID数据保存成功, 共{len(comics)}本漫画{Fore.RESET}")
    elif args.id:
        if cl.confirm():
            comics = cl.get_comics_details(args.id)
            dm.save(comics)
            tqdm.write(f"{Fore.GREEN}自定义漫画ID数据保存成功, 共{len(comics)}本漫画{Fore.RESET}")
    elif args.type == 'classify':
        if cl.confirm():
            comics = cl.get_classify_page_all()
            dm.save(comics)
            tqdm.write(f"{Fore.GREEN}[分类页]数据保存成功, 共{len(comics)}本漫画{Fore.RESET}")
    elif args.type == 'ranking':
        if cl.confirm():
            page = cl.get_ranking_page(args.rank)
            comic_id_list = [i.get("comic_id") for i in page.get("rankListInfo")]
            comics = cl.get_ranking_page(comic_id_list).get("rankListInfo", {})
            for index, comic in enumerate(comics):
                comic["rank"] = index + 1
            dm.save(comics)
            tqdm.write(f"{Fore.GREEN}[{cl.ranking_dict[args.rank]}]数据保存成功, 共{len(comics)}本漫画{Fore.RESET}")
    elif args.type == 'update':
        if cl.confirm():
            comics = cl.get_update_page_all()
            dm.save(comics)
            tqdm.write(f"{Fore.GREEN}[更新推荐页]数据保存成功, 共{len(comics)}本漫画{Fore.RESET}")
    elif args.type == 'home_feed':
        if cl.confirm():
            comics = cl.get_home_feeds_all()
            dm.save(comics)
            tqdm.write(f"{Fore.GREEN}[主页信息流]数据保存成功, 共{len(comics)}本漫画{Fore.RESET}")
    elif args.type == 'favorite':
        if not args.headers:
            print(f"{Fore.RED}请使用参数--headers导入正确填写Cookie的json文件{Fore.RESET}")
            return
        if cl.confirm():
            comics = cl.get_favorite_all()
            dm.save(comics)
            tqdm.write(f"{Fore.GREEN}[我的追漫]数据保存成功, 共{len(comics)}本漫画{Fore.RESET}")
    elif args.type == 'buy':
        if cl.confirm():
            comics = cl.get_buy_comics()
            dm.save(comics)
            tqdm.write(f"{Fore.GREEN}[已购漫画]数据保存成功, 共{len(comics)}本漫画{Fore.RESET}")

    if not args.id and args.detail:
        args.id = [comic['comic_id'] for comic in comics]
        comics = cl.get_comics_details(args.id)
        dm.save(comics)
        tqdm.write(f"{Fore.GREEN}漫画详情页保存成功, 共{len(comics)}本漫画{Fore.RESET}")

    if args.bonus:
        if cl.get_comic_bonus_all(comics):
            dm.save(comics)
            tqdm.write(f"{Fore.GREEN}特典数据保存成功{Fore.RESET}")

if __name__ == "__main__":
    if is_launched_by_explorer():
        run_gui()
    else:
        args = parse_args()
        crawler = Crawler(args)
        document = Document(args)
        try:
            run_cli(args, crawler, document)
        except KeyboardInterrupt:
            pass
        except Exception:
            print(f"{Fore.RED}程序执行出现错误, 意外退出\n{traceback.print_exc()}{Fore.RESET}")
